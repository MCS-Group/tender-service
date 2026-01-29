import openpyxl
from openpyxl.utils import get_column_letter    
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from typing import List
from datetime import datetime
from schemas.lvl_schema import TenderOverview, dict_of_level, BudgetType
from src.logger import logger

def save_to_excel(overviews: List[dict], filename: str):
    logger.info(f"Saving {len(overviews)} tender overviews to Excel: {filename}")
    wb = openpyxl.Workbook()
    
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()

    ws.title = "Tender Report"


    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Dark Blue
    header_font = Font(color="FFFFFF", bold=True, name="Times New Roman", size=10)
    data_font = Font(name="Times New Roman", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_align_top = Alignment(horizontal="center", vertical="top", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Main Header "Тендерийн бүртгэл"
    ws.merge_cells("A1:P1")  # Span all data columns (A–P)
    ws["A1"] = "Тендерийн бүртгэл"
    ws["A1"].font = Font(bold=True, size=14, name="Times New Roman")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    # Date Row
    ws.merge_cells("A2:P2")
    current_date = datetime.now().strftime("%Y.%m.%d")
    ws["A2"] = f"Тайлант хугацаа: {current_date}"
    ws["A2"].font = Font(italic=True, size=10, name="Times New Roman")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Table Header "ЗАРЛАСАН ТЕНДЕРИЙН МЭДЭЭЛЭЛ"
    ws.merge_cells("A3:P3")
    ws["A3"] = "ЗАРЛАСАН ТЕНДЕРИЙН МЭДЭЭЛЭЛ"
    ws["A3"].fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid") # Lighter Blue
    ws["A3"].font = header_font
    ws["A3"].alignment = center_align
    # Apply border to merged header row
    for row in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=16):
        for cell in row:
            cell.border = thin_border

    # Column Headers
    headers = [
        "№", 
        "Summary",
        "Category",
        "Category Detail",
        "Тендерийн нэр", 
        "Захиалагч", 
        "Нийт төсөвт өртөг", 
        "Budget Type",
        "Зарласан огноо", 
        "Хүлээн авах огноо", 
        "Шалгаруулалтын дугаар", 
        "Link",
        "Tender Type",
        "Level 1",
        "Level 2",
        "Level 3"
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Adjust column widths (A–P)
    column_widths = [
        5,   # №
        45,  # Summary
        18,  # Category
        22,  # Category Detail
        35,  # Тендерийн нэр
        25,  # Захиалагч
        18,  # Нийт төсөвт өртөг
        16,  # Budget Type
        12,  # Зарласан огноо
        12,  # Хүлээн авах огноо
        20,  # Шалгаруулалтын дугаар
        35,  # Link
        10,  # Tender Type
        18,  # Level 1
        22,  # Level 2
        22,  # Level 3
    ]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Give header rows a bit more height so they never clip
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[3].height = 20

    def normalize_list(value):
        if value is None:
            return []
        if isinstance(value, list):
            items = value
        else:
            items = [value]

        normalized = []
        seen = set()
        for item in items:
            if item is None:
                continue
            text = str(item).strip()
            if not text:
                continue
            key = text.casefold()
            if key in seen:
                continue
            seen.add(key)
            normalized.append(text)
        return normalized

    def infer_budget_type(overview: dict) -> str:
        raw = (
            overview.get("budget_type")
            or overview.get("budgetType")
            or overview.get("budget_funding")
            or overview.get("budgetFunding")
        )
        if isinstance(raw, BudgetType):
            return raw.value
        if isinstance(raw, str) and raw.strip():
            return raw.strip()

        text = f"{overview.get('name', '')} {overview.get('summary', '')}".casefold()

        if "орон нутгийн төсөв" in text:
            return BudgetType.LOCAL_BUDGET.value
        if "урсгал төсөв" in text:
            return BudgetType.CURRENT_BUDGET.value
        if "багц санхүүжилт" in text or "багц-" in text:
            return BudgetType.PACKAGE_BUDGET.value
        if "өөрийн хөрөнгө" in text:
            return BudgetType.OWN_FUNDS.value
        if "төсвийн хөрөнгө" in text:
            return BudgetType.STATE_BUDGET.value

        loan_aid_markers = (
            "дэлхийн банк",
            "world bank",
            "ахб",
            "азийн хөгжлийн банк",
            "asian development bank",
            "jica",
            "kfw",
            "ebrd",
            "eib",
            "usaid",
            "undp",
            "unicef",
            "who",
            "ifad",
            "eu",
            "european union",
            "зээл",
            "тусламж",
        )
        if any(m in text for m in loan_aid_markers) and "санхүүжил" in text:
            return BudgetType.LOAN_AID_FUNDING.value

        return ""

    def pick_from_list(items: List[str], idx: int, repeat_single: bool = True) -> str:
        if idx < len(items):
            return items[idx]
        if repeat_single and len(items) == 1:
            return items[0]
        return ""

    def _estimate_line_count(value: object, col_width: float) -> int:
        if value is None:
            return 1
        if isinstance(value, (int, float)):
            return 1
        text = str(value)
        if not text:
            return 1
        # Roughly map Excel column width -> characters per line.
        # This is heuristic; the goal is to avoid clipping, not perfect pixel matching.
        chars_per_line = max(8, int(col_width * 1.15))
        total_lines = 0
        for seg in text.splitlines() or [text]:
            seg = seg.strip("\r")
            seg_len = len(seg)
            total_lines += max(1, (seg_len + chars_per_line - 1) // chars_per_line)
        return max(1, total_lines)

    def _ensure_group_height(
        start_row: int,
        end_row: int,
        col_idx: int,
        value: object,
        base_row_height: float = 15.0,
        max_row_height: float = 240.0,
    ) -> None:
        """Ensure total height across rows in [start_row..end_row] can show wrapped text."""
        needed_lines = _estimate_line_count(value, column_widths[col_idx - 1])
        needed_total = min(max_row_height * (end_row - start_row + 1), base_row_height * needed_lines)
        current_total = 0.0
        for r in range(start_row, end_row + 1):
            current_total += float(ws.row_dimensions[r].height or base_row_height)
        if current_total >= needed_total:
            return
        extra = needed_total - current_total
        per_row_extra = extra / max(1, (end_row - start_row + 1))
        for r in range(start_row, end_row + 1):
            cur = float(ws.row_dimensions[r].height or base_row_height)
            ws.row_dimensions[r].height = min(max_row_height, cur + per_row_extra)

    # Helper function to convert lists to strings and map codes to descriptions
    def format_value(value, use_dict=False, sep=", "):
        if isinstance(value, list):
            if use_dict:
                descriptions = [dict_of_level.get(str(v), str(v)) for v in value]
                return sep.join(str(v) for v in normalize_list(descriptions))
            return sep.join(str(v) for v in normalize_list(value))
        if use_dict and value:
            return dict_of_level.get(str(value), str(value))
        return value if value is not None else ""

    # Data Rows (grouped): each category/detail is its own row; Summary is merged vertically
    data_row = 5
    tender_no = 1
    merge_cols = [1, 2, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]

    base_row_height = 15.0
    max_row_height = 240.0

    for overview in overviews:
        categories = normalize_list(overview.get("tender_category", []))
        details = normalize_list(overview.get("tender_category_detail", []))
        level2_list = normalize_list(overview.get("level2", []))
        level3_list = normalize_list(overview.get("level3", []))

        group_len = max(len(categories), len(details), len(level2_list), len(level3_list), 1)
        start_row = data_row
        end_row = data_row + group_len - 1

        # Default each row in the group to base height; we'll increase as needed.
        for r in range(start_row, end_row + 1):
            if ws.row_dimensions[r].height is None:
                ws.row_dimensions[r].height = base_row_height

        for i in range(group_len):
            r = data_row + i

            row_data = [
                tender_no if i == 0 else None,  # №
                format_value(overview.get("summary", "")) if i == 0 else None,
                format_value(pick_from_list(categories, i), use_dict=True),
                format_value(pick_from_list(details, i), use_dict=True),
                format_value(overview.get("name", "")) if i == 0 else None,
                format_value(overview.get("ordering_organization", "")) if i == 0 else None,
                format_value(overview.get("total_budget", "")) if i == 0 else None,
                infer_budget_type(overview) if i == 0 else None,
                format_value(overview.get("announced_date", "")) if i == 0 else None,
                format_value(overview.get("deadline_date", "")) if i == 0 else None,
                format_value(overview.get("selection_number", "")) if i == 0 else None,
                format_value(overview.get("official_link", "")) if i == 0 else None,
                format_value(overview.get("tender_type", ""), use_dict=True) if i == 0 else None,
                format_value(overview.get("level1", "")) if i == 0 else None,
                format_value(pick_from_list(level2_list, i)),
                format_value(pick_from_list(level3_list, i)),
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = left_align

                if col_idx == 1:
                    cell.alignment = center_align_top

                # Category/Detail/Level2/Level3 look better centered like the mockup
                if col_idx in (3, 4, 15, 16):
                    cell.alignment = center_align_top

                if col_idx == 7 and i == 0 and value is not None:
                    cell.number_format = '#,##0.00'

                if col_idx == 12 and i == 0:
                    cell.font = Font(name="Times New Roman", size=10, color="0000FF", underline="single")
                    if value:
                        cell.hyperlink = value

                # Per-row auto height for non-merged, wrapped cells
                if value not in (None, ""):
                    needed_lines = _estimate_line_count(value, column_widths[col_idx - 1])
                    needed_height = min(max_row_height, base_row_height * needed_lines)
                    cur_h = float(ws.row_dimensions[r].height or base_row_height)
                    if needed_height > cur_h:
                        ws.row_dimensions[r].height = needed_height

        if group_len > 1:
            for col_idx in merge_cols:
                ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)
                # Apply border to all cells inside merged ranges
                for rr in range(start_row, end_row + 1):
                    ws.cell(row=rr, column=col_idx).border = thin_border

            # Ensure merged cells have enough total height across the group.
            _ensure_group_height(start_row, end_row, 2, overview.get("summary", ""), base_row_height, max_row_height)
            _ensure_group_height(start_row, end_row, 5, overview.get("name", ""), base_row_height, max_row_height)
            _ensure_group_height(start_row, end_row, 6, overview.get("ordering_organization", ""), base_row_height, max_row_height)
            _ensure_group_height(start_row, end_row, 12, overview.get("official_link", ""), base_row_height, max_row_height)

            # If there is a single category/level2 but multiple detail rows, merge Category/Level2 too
            if len(categories) == 1:
                ws.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)
                for rr in range(start_row, end_row + 1):
                    ws.cell(row=rr, column=3).border = thin_border
            if len(level2_list) == 1:
                ws.merge_cells(start_row=start_row, start_column=15, end_row=end_row, end_column=15)
                for rr in range(start_row, end_row + 1):
                    ws.cell(row=rr, column=15).border = thin_border

        data_row += group_len
        tender_no += 1

    # Freeze panes so headers remain visible while scrolling
    ws.freeze_panes = "A5"


    try:
        wb.save(filename)
        logger.info(f"Excel file saved successfully: {filename}")
    except PermissionError:
        logger.error(f"Permission denied: Cannot save '{filename}'. Please close the file if it's open.")
        raise
    except Exception as e:
        logger.error(f"Error saving Excel file '{filename}': {e}", exc_info=True)
        raise


def pdf_result_to_excel(raw_overviews: List[dict], filename: str):
    # Filter parts: if main_category is "Нарийн хүнсний ногоо нийлүүлэл", 
    # only keep parts where food_category is "Target Vegetable"
    overviews = []

    for overview in raw_overviews:
        parts = overview.get("parts", [])
        main_category = overview.get("main_category", "")
        
        if main_category == "Нарийн хүнсний ногоо нийлүүлэл":
            if overview.get("is_allowed"):
                if parts:
                    filter_parts = []
                    for part in parts:
                        if part.get("food_category") == "Target Vegetable":
                            filter_parts.append(part)
                    if filter_parts:
                        overview["parts"] = filter_parts
                        overviews.append(overview)
                else:
                    overviews.append(overview)
        else:
            overviews.append(overview)





    wb = openpyxl.Workbook()
    
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()

    ws.title = "Tender Report"

    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Dark Blue
    header_font = Font(color="FFFFFF", bold=True, name="Times New Roman", size=10)
    data_font = Font(name="Times New Roman", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_align_top = Alignment(horizontal="center", vertical="top", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Main Header
    ws.merge_cells("A1:O1")
    ws["A1"] = "Тендерийн бүртгэл"
    ws["A1"].font = Font(bold=True, size=14, name="Times New Roman")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    # Date Row
    ws.merge_cells("A2:O2")
    current_date = datetime.now().strftime("%Y.%m.%d")
    ws["A2"] = f"Тайлант хугацаа: {current_date}"
    ws["A2"].font = Font(italic=True, size=10, name="Times New Roman")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Table Header
    ws.merge_cells("A3:O3")
    ws["A3"] = "ЗАРЛАСАН ТЕНДЕРИЙН ДЭЛГЭРЭНГҮЙ МЭДЭЭЛЭЛ"
    ws["A3"].fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid") # Lighter Blue
    ws["A3"].font = header_font
    ws["A3"].alignment = center_align
    for row in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=15):
        for cell in row:
            cell.border = thin_border

    # Column Headers
    headers = [
        "№", "Main Category", "Тендерийн нэр", "Захиалагч", "Нийт төсөвт өртөг", "Зарласан огноо", 
        "Хүлээн авах огноо", "Шалгаруулалтын дугаар", "Link",
        "Part Name", "Part Content", "Part Budget",
        "Main Requirements", "Business Requirements", "Technical Requirements"
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Adjust column widths
    column_widths = [
        5, 20, 35, 25, 18, 12, 12, 20, 35,
        30, 45, 18,
        45, 45, 45
    ]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[3].height = 20

    def format_value(value, sep=", "):
        if isinstance(value, list):
            return sep.join(map(str, value))
        return value if value is not None else ""

    def estimate_row_height(text, col_width, base_height=15.0, max_height=300.0):
        """Estimate required row height based on text length and column width"""
        if not text:
            return base_height
        text_str = str(text)
        # Approximate characters per line based on column width
        chars_per_line = max(8, int(col_width * 1.2))
        # Count lines needed
        lines_needed = 1
        for line in text_str.split('\n'):
            line_len = len(line)
            lines_needed += max(1, (line_len + chars_per_line - 1) // chars_per_line)
        # Calculate height (roughly 15 points per line)
        estimated_height = min(max_height, base_height * lines_needed)
        return estimated_height

    data_row = 5
    tender_no = 1
    base_row_height = 15.0

    for overview in overviews:
        parts = overview.get("parts") or []
        requirements = overview.get("requirements") or {}
        

        group_len = max(len(parts), 1)
        start_row = data_row
        end_row = data_row + group_len - 1

        for r in range(start_row, end_row + 1):
            if ws.row_dimensions[r].height is None:
                ws.row_dimensions[r].height = base_row_height

        # Merged tender-level info
        tender_info_cols = {
            1: tender_no,
            2: overview.get("main_category", ""),
            3: overview.get("name", ""),
            4: overview.get("ordering_organization", ""),
            5: overview.get("total_budget", ""),
            6: overview.get("announced_date", ""),
            7: overview.get("deadline_date", ""),
            8: overview.get("selection_number", ""),
            9: overview.get("official_link", ""),
            13: requirements.get("main_requirements", ""),
            14: requirements.get("business_requirements", ""),
            15: requirements.get("technical_requirements", ""),
        }

        for col_idx, value in tender_info_cols.items():
            cell = ws.cell(row=start_row, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = left_align if col_idx not in (1, 2) else center_align_top
            if col_idx == 5 and value:
                cell.number_format = '#,##0.00'
            if col_idx == 9 and value:
                cell.font = Font(name="Times New Roman", size=10, color="0000FF", underline="single")
                cell.hyperlink = value

        for i in range(group_len):
            r = data_row + i
            
            part = parts[i] if i < len(parts) else {}

            # Part-level info
            part_cols = {
                10: part.get("part_name", ""),
                11: part.get("content", ""),
                12: part.get("part_budget", ""),
            }
            for col_idx, value in part_cols.items():
                cell = ws.cell(row=r, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = left_align
                if col_idx == 12 and value:
                    cell.number_format = '#,##0.00'

        # Apply merges
        if group_len > 1:
            for col_idx in tender_info_cols.keys():
                 ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)
                 for rr in range(start_row, end_row + 1):
                    ws.cell(row=rr, column=col_idx).border = thin_border

        # Auto-adjust row heights for long content in requirements columns
        max_height_needed = base_row_height
        for col_idx in [3, 4, 10, 11, 13, 14, 15]:  # Name, Organization, Part Name, Part Content, Requirements columns
            if col_idx in tender_info_cols:
                value = tender_info_cols[col_idx]
                if value:
                    height = estimate_row_height(value, column_widths[col_idx - 1])
                    max_height_needed = max(max_height_needed, height)
        
        # Check part content heights
        for i in range(group_len):
            part = parts[i] if i < len(parts) else {}
            for col_idx in [10, 11]:  # Part Name, Part Content
                value = part.get("part_name" if col_idx == 10 else "content", "")
                if value:
                    height = estimate_row_height(value, column_widths[col_idx - 1])
                    if height > ws.row_dimensions[data_row + i].height:
                        ws.row_dimensions[data_row + i].height = height
        
        # Apply calculated height to merged rows
        if group_len > 1:
            height_per_row = max_height_needed / group_len
            for r in range(start_row, end_row + 1):
                current = ws.row_dimensions[r].height or base_row_height
                ws.row_dimensions[r].height = max(current, height_per_row)
        else:
            ws.row_dimensions[start_row].height = max(ws.row_dimensions[start_row].height, max_height_needed)

        data_row += group_len
        tender_no += 1

    ws.freeze_panes = "A5"

    try:
        wb.save(filename)
    except PermissionError:
        print(f"Error: Permission denied. Close the file '{filename}' if it's open.")
